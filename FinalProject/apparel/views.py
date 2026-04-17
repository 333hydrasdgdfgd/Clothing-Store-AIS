from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum
from django.http import JsonResponse
from .models import (
    Product, Category, Brand, Size, Color, Material,
    Supplier, Customer, Employee, PurchaseOrder, PurchaseOrderDetails,
    SalesOrder, SalesOrderDetails, ShippingDetails, Invoice, Receipt, Expense,
    AuditLog,
)
from datetime import timedelta

def audit(request, action, description):
    emp_id   = request.session.get('employee_id')
    emp_name = request.session.get('employee_name', 'Unknown')
    employee = None
    if emp_id:
        try:
            employee = Employee.objects.get(pk=emp_id)
        except Employee.DoesNotExist:
            pass
    AuditLog.objects.create(
        employee=employee,
        employee_name=emp_name,
        action=action,
        description=description,
    )


def coverpage(request):
    return render(request, 'apparel/coverpage.html')

def dashboard(request):
    employee_id = request.session.get('employee_id')
    if not employee_id:
        return redirect('login')

    role = request.session.get('employee_role')
    access = {
        'products': False,
        'inventory': False,
        'suppliers': False,
        'purchase_orders': False,
        'sales_orders': False,
        'customers': False,
        'employees': False,
        'shipping': False,
        'accounting': False,
        'audit_log': False,
    }

    if role == 'Admin':
        for key in access: access[key] = True
    elif role == 'Purchasing':
        access.update({'products': True, 'suppliers': True, 'purchase_orders': True})
    elif role == 'Inventory':
        access.update({'inventory': True, 'purchase_orders': True, 'shipping': True})
    elif role == 'Sales':
        access.update({'products': True, 'sales_orders': True, 'customers': True})
    elif role in ('Accounting Manager', 'Staff Accountant'):
        access.update({'accounting': True})

    total_revenue = SalesOrder.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    total_sales_orders = SalesOrder.objects.count()
    total_purchase_orders = PurchaseOrder.objects.count()
    total_products = Product.objects.count()
    total_suppliers = Supplier.objects.count()
    total_customers = Customer.objects.count()
    pending_pos = PurchaseOrder.objects.filter(status='Pending').count()
    recent_sales = SalesOrder.objects.select_related('customer', 'employee').order_by('-sales_id')[:5]
    recent_pos = PurchaseOrder.objects.select_related('supplier', 'employee').order_by('-purchase_order_id')[:5]

    # ── Dashboard graph data: monthly revenue & orders (last 6 months) ──
    import json
    from datetime import date
    today = date.today()
    months_labels = []
    monthly_revenue = []
    monthly_orders = []
    for i in range(5, -1, -1):
        # Go back i months
        month_date = (today.replace(day=1) - timedelta(days=1)) if i == 0 else today
        # Simpler: compute year/month directly
        month_num = (today.month - i - 1) % 12 + 1
        year_offset = (today.month - i - 1) // 12
        year_num = today.year - year_offset
        # Correct calculation
        total_months = today.month - 1 - i  # 0-indexed month offset from Jan of this year
        year_num = today.year + total_months // 12
        month_num = total_months % 12 + 1
        if month_num <= 0:
            month_num += 12
            year_num -= 1

        rev = SalesOrder.objects.filter(
            sales_date__year=year_num, sales_date__month=month_num
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        orders_count = SalesOrder.objects.filter(
            sales_date__year=year_num, sales_date__month=month_num
        ).count()

        import calendar
        months_labels.append(calendar.month_abbr[month_num])
        monthly_revenue.append(float(rev))
        monthly_orders.append(orders_count)

    context = {
        'employee_name': request.session.get('employee_name'),
        'employee_role': role,
        'access': access,
        'total_revenue': total_revenue,
        'total_sales_orders': total_sales_orders,
        'total_purchase_orders': total_purchase_orders,
        'total_products': total_products,
        'total_suppliers': total_suppliers,
        'total_customers': total_customers,
        'pending_pos': pending_pos,
        'recent_sales': recent_sales,
        'recent_pos': recent_pos,
        # graph data
        'chart_labels_json': json.dumps(months_labels),
        'chart_revenue_json': json.dumps(monthly_revenue),
        'chart_orders_json': json.dumps(monthly_orders),
    }

    return render(request, 'apparel/dashboard.html', context)


def login_view(request):
    import time
    error = None
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        attempts = request.session.get('login_attempts', {})
        user_data = attempts.get(username, {'count': 0, 'lockout_until': None})

        lockout_until = user_data.get('lockout_until')
        current_time = time.time()
        if lockout_until and current_time < lockout_until:
            remaining = int(lockout_until - current_time) + 1
            error = f"Too many failed attempts. Try again in {remaining} seconds."
        else:
            try:
                from .models import Employee as EmployeeModel
                emp = EmployeeModel.objects.get(username=username)
                if emp.check_password(password):
                    attempts.pop(username, None)
                    request.session['login_attempts'] = attempts
                    request.session['employee_id'] = emp.id
                    request.session['employee_name'] = emp.full_name
                    request.session['employee_role'] = emp.role
                    # ── AUDIT: login ──
                    AuditLog.objects.create(
                        employee=emp,
                        employee_name=emp.full_name,
                        action='LOGIN',
                        description=f'{emp.full_name} logged in.',
                    )
                    return redirect('dashboard')
                else:
                    raise ValueError("bad password")
            except Exception:
                user_data['count'] = user_data.get('count', 0) + 1
                if user_data['count'] >= 3:
                    user_data['lockout_until'] = current_time + 30
                    error = "Too many failed attempts. Please wait 30 seconds."
                else:
                    remaining_attempts = 3 - user_data['count']
                    error = f"Invalid username or password. {remaining_attempts} attempt(s) remaining."
                attempts[username] = user_data
                request.session['login_attempts'] = attempts

    return render(request, 'apparel/login.html', {'error': error})

def logout_view(request):
    # ── AUDIT: logout ──
    emp_id   = request.session.get('employee_id')
    emp_name = request.session.get('employee_name', 'Unknown')
    if emp_id:
        try:
            emp = Employee.objects.get(pk=emp_id)
            AuditLog.objects.create(
                employee=emp,
                employee_name=emp_name,
                action='LOGOUT',
                description=f'{emp_name} logged out.',
            )
        except Employee.DoesNotExist:
            pass
    request.session.flush()
    return redirect('login')

# Product List
def product_list(request):
    query = request.GET.get('q')
    products = Product.objects.prefetch_related('sizes', 'colors').all()
    if query:
        products = products.filter(product_name__icontains=query)
    return render(request, 'apparel/product_list.html', {'products': products, 'query': query})

# Add Product
def product_create(request):
    categories = Category.objects.all()
    brands = Brand.objects.all()
    sizes = Size.objects.all()
    colors = Color.objects.all()
    materials = Material.objects.all()

    if request.method == 'POST':
        size_ids = request.POST.getlist('sizes[]')
        color_ids = request.POST.getlist('colors[]')

        product = Product.objects.create(
            product_code=request.POST['product_code'],
            product_name=request.POST['product_name'],
            category=Category.objects.get(pk=request.POST['category']),
            brand=Brand.objects.get(pk=request.POST['brand']),
            material=Material.objects.get(pk=request.POST['material']),
            unit_cost=request.POST['unit_cost'],
            selling_price=request.POST['selling_price']
        )

        if size_ids:
            product.sizes.set(size_ids)
        if color_ids:
            product.colors.set(color_ids)

        audit(request, 'CREATE', f'{request.session.get("employee_name")} created product "{product.product_name}" (Code: {product.product_code}).')
        return redirect('product_list')

    context = {
        'categories': categories,
        'brands': brands,
        'sizes': sizes,
        'colors': colors,
        'materials': materials
    }
    return render(request, 'apparel/add_product.html', context)

def receive_order(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    if request.method == 'POST' and po.status != 'Received':
        from django.utils import timezone as tz
        po.status = 'Received'
        po.received_date = tz.localtime(tz.now())
        po.save()

        for detail in po.details.select_related('product').all():
            product = detail.product
            old_cost = float(product.unit_cost)
            new_cost = float(detail.unit_cost)
            if new_cost > old_cost:
                if old_cost > 0:
                    markup_ratio = float(product.selling_price) / old_cost
                    product.selling_price = round(new_cost * markup_ratio, 2)
                product.unit_cost = new_cost
                product.save()

        try:
            _ = po.receipt
        except Exception:
            Receipt.objects.create(
                ref=f'REC-{po.purchase_order_id}',
                date=tz.now().date(),
                source=po.supplier.supplier_name,
                description=f'Purchase Order PO-{po.purchase_order_id} received',
                amount=po.total_cost,
                receipt_type='purchase',
                purchase_order=po,
            )

        audit(request, 'RECEIVE', f'{request.session.get("employee_name")} received Purchase Order PO-{po.purchase_order_id} from {po.supplier.supplier_name}.')

    return redirect('purchase_order_details', pk=pk)

# Edit Product
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    categories = Category.objects.all()
    brands = Brand.objects.all()
    sizes = Size.objects.all()
    colors = Color.objects.all()
    materials = Material.objects.all()

    if request.method == 'POST':
        size_ids = request.POST.getlist('sizes')
        color_ids = request.POST.getlist('colors')

        product.product_code = request.POST['product_code']
        product.product_name = request.POST['product_name']
        product.category = Category.objects.get(pk=request.POST['category'])
        product.brand = Brand.objects.get(pk=request.POST['brand'])
        product.material = Material.objects.get(pk=request.POST['material'])
        product.unit_cost = request.POST['unit_cost']
        product.selling_price = request.POST['selling_price']
        product.save()

        product.sizes.set(size_ids)
        product.colors.set(color_ids)

        audit(request, 'EDIT', f'{request.session.get("employee_name")} edited product "{product.product_name}" (ID: {product.pk}).')
        return redirect('product_list')

    context = {
        'product': product,
        'categories': categories,
        'brands': brands,
        'sizes': sizes,
        'colors': colors,
        'materials': materials
    }
    return render(request, 'apparel/edit_product.html', context)

# Delete Product
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    name = product.product_name
    product.delete()
    audit(request, 'DELETE', f'{request.session.get("employee_name")} deleted product "{name}" (ID: {pk}).')
    return redirect('product_list')

def get_products_by_supplier(request, supplier_id):
    products = Product.objects.filter(brand__supplier_id=supplier_id).values('id', 'product_name', 'product_code', 'unit_cost')
    return JsonResponse({'products': list(products)})

# JSON endpoint to fetch attributes by category
def get_attributes(request, category_id):
    brands = list(Brand.objects.filter(category_id=category_id).values('id', 'brand_name'))
    sizes = list(Size.objects.filter(category_id=category_id).values('id', 'size_name'))
    colors = list(Color.objects.filter(category_id=category_id).values('id', 'color_name'))
    materials = list(Material.objects.filter(category_id=category_id).values('id', 'material_name'))

    return JsonResponse({
        'brands': brands,
        'sizes': sizes,
        'colors': colors,
        'materials': materials,
    })

# ── NEW: JSON endpoint — all products for a given category (used in sales order JS) ──
def get_products_by_category(request, category_id):
    products = Product.objects.filter(category_id=category_id).prefetch_related('sizes', 'colors')
    data = []
    for p in products:
        data.append({
            'id': p.id,
            'product_name': p.product_name,
            'product_code': p.product_code,
            'selling_price': float(p.selling_price),
            'brand_id': p.brand_id,
            'colors': [{'id': c.id, 'name': c.color_name} for c in p.colors.all()],
            'sizes':  [{'id': s.id, 'name': s.size_name}  for s in p.sizes.all()],
        })
    return JsonResponse({'products': data})


# Supplier List
def supplier_list(request):
    query = request.GET.get('q')
    suppliers = Supplier.objects.all()
    if query:
        suppliers = suppliers.filter(supplier_name__icontains=query)
    return render(request, 'apparel/supplier_list.html', {'suppliers': suppliers, 'query': query})

# Add Supplier
def supplier_create(request):
    if request.method == 'POST':
        s = Supplier.objects.create(
            supplier_name=request.POST['supplier_name'],
            contact_person=request.POST.get('contact_person', ''),
            phone=request.POST.get('phone', ''),
            email=request.POST.get('email', ''),
            address=request.POST.get('address', '')
        )
        audit(request, 'CREATE', f'{request.session.get("employee_name")} created supplier "{s.supplier_name}".')
        return redirect('supplier_list')
    return render(request, 'apparel/add_supplier.html')

# Edit Supplier
def supplier_update(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.supplier_name = request.POST['supplier_name']
        supplier.contact_person = request.POST.get('contact_person', '')
        supplier.phone = request.POST.get('phone', '')
        supplier.email = request.POST.get('email', '')
        supplier.address = request.POST.get('address', '')
        supplier.save()
        audit(request, 'EDIT', f'{request.session.get("employee_name")} edited supplier "{supplier.supplier_name}" (ID: {pk}).')
        return redirect('supplier_list')
    return render(request, 'apparel/edit_supplier.html', {'supplier': supplier})

# Delete Supplier
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    name = supplier.supplier_name
    supplier.delete()
    audit(request, 'DELETE', f'{request.session.get("employee_name")} deleted supplier "{name}" (ID: {pk}).')
    return redirect('supplier_list')

# Customer list
def customer_list(request):
    query = request.GET.get('q')
    customers = Customer.objects.all()
    if query:
        customers = customers.filter(full_name__icontains=query)
    return render(request, 'apparel/customer_list.html', {'customers': customers, 'query': query})

# Add Customer
def customer_create(request):
    next_url = request.GET.get('next') or request.POST.get('next') or ''
    if request.method == 'POST':
        c = Customer.objects.create(
            full_name=request.POST['full_name'],
            email=request.POST.get('email', ''),
            phone=request.POST.get('phone', ''),
            address=request.POST.get('address', '')
        )
        audit(request, 'CREATE', f'{request.session.get("employee_name")} created customer "{c.full_name}".')
        if next_url:
            return redirect(next_url)
        return redirect('customer_list')
    return render(request, 'apparel/add_customer.html', {'next': next_url})

# Edit Customer
def customer_update(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.full_name = request.POST['full_name']
        customer.email = request.POST.get('email', '')
        customer.phone = request.POST.get('phone', '')
        customer.address = request.POST.get('address', '')
        customer.save()
        audit(request, 'EDIT', f'{request.session.get("employee_name")} edited customer "{customer.full_name}" (ID: {pk}).')
        return redirect('customer_list')
    return render(request, 'apparel/edit_customer.html', {'customer': customer})

# Delete Customer
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    name = customer.full_name
    customer.delete()
    audit(request, 'DELETE', f'{request.session.get("employee_name")} deleted customer "{name}" (ID: {pk}).')
    return redirect('customer_list')

# Employee list
def employee_list(request):
    query = request.GET.get('q')
    employees = Employee.objects.all().order_by('id')
    if query:
        employees = employees.filter(full_name__icontains=query).order_by('id')
    return render(request, 'apparel/employee_list.html', {'employees': employees, 'query': query})

# Add Employee
def employee_create(request):
    if request.method == 'POST':
        from django.contrib.auth.hashers import make_password
        raw_password = request.POST['password']
        employee = Employee(
            full_name=request.POST['full_name'],
            username=request.POST['username'],
            role=request.POST['role'],
            email=request.POST.get('email', ''),
            phone=request.POST.get('phone', ''),
            address=request.POST['address'],
            salary=request.POST['salary']
        )
        employee.password = make_password(raw_password)
        employee.save()
        audit(request, 'CREATE', f'{request.session.get("employee_name")} created employee "{employee.full_name}" with role {employee.role}.')
        return redirect('employee_list')
    return render(request, 'apparel/add_employee.html')

# Edit Employee
def employee_update(request, pk):
    employee = Employee.objects.get(pk=pk)
    if request.method == "POST":
        from django.contrib.auth.hashers import make_password
        employee.full_name = request.POST.get('full_name')
        employee.username = request.POST.get('username')
        employee.role = request.POST.get('role')
        employee.email = request.POST.get('email')
        employee.phone = request.POST.get('phone')
        employee.address = request.POST.get('address')
        employee.salary = request.POST.get('salary')
        raw_password = request.POST.get('password')
        if raw_password:
            employee.password = make_password(raw_password)
        employee.save()
        audit(request, 'EDIT', f'{request.session.get("employee_name")} edited employee "{employee.full_name}" (ID: {pk}).')
        return redirect('employee_list')
    return render(request, 'apparel/edit_employee.html', {'employee': employee})

# Delete Employee
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    name = employee.full_name
    employee.delete()
    audit(request, 'DELETE', f'{request.session.get("employee_name")} deleted employee "{name}" (ID: {pk}).')
    return redirect('employee_list')

# Purchase Order
def purchase_orders(request):
    purchase_orders = PurchaseOrder.objects.select_related('supplier', 'employee').all()
    return render(request, 'apparel/purchase_order_list.html', {'purchase_orders': purchase_orders})


def create_purchase_order(request):
    suppliers = Supplier.objects.all()
    products = Product.objects.all()

    if request.method == "POST":
        supplier_id = request.POST.get("supplier")
        expected_delivery = request.POST.get("expected_delivery")
        po_discount = request.POST.get("po_discount") or 0
        status = request.POST.get("status") or "Pending"

        supplier = get_object_or_404(Supplier, id=supplier_id)
        session_employee_id = request.session.get('employee_id')
        employee = Employee.objects.get(id=session_employee_id) if session_employee_id else None

        purchase_order = PurchaseOrder.objects.create(
            supplier=supplier,
            employee=employee,
            expected_delivery_date=expected_delivery,
            total_cost=0,
            po_discount=po_discount,
            status=status
        )

        product_ids = request.POST.getlist("product[]")
        quantities = request.POST.getlist("quantity[]")
        costs = request.POST.getlist("cost[]")

        total = 0

        for i in range(len(product_ids)):
            product = get_object_or_404(Product, id=product_ids[i])
            quantity = int(quantities[i])
            unit_cost = float(costs[i])
            subtotal = quantity * unit_cost
            total += subtotal

            PurchaseOrderDetails.objects.create(
                purchase_order=purchase_order,
                product=product,
                quantity_ordered=quantity,
                unit_cost=unit_cost,
                subtotal=subtotal
            )

        discount_amount = total * (float(po_discount) / 100)
        final_total = total - discount_amount

        purchase_order.total_cost = final_total
        purchase_order.save()

        audit(request, 'CREATE', f'{request.session.get("employee_name")} created Purchase Order PO-{purchase_order.purchase_order_id} for supplier "{supplier.supplier_name}".')
        return redirect('purchase_orders')

    return render(request, 'apparel/create_purchase_order.html', {
        'suppliers': suppliers,
        'products': products,
        'employee_name': request.session.get('employee_name', 'Unknown'),
    })

def purchase_order_details(request, pk):
    purchase_order = get_object_or_404(PurchaseOrder, pk=pk)
    po_details = PurchaseOrderDetails.objects.filter(
        purchase_order=purchase_order
    ).select_related('product')
    return render(request, 'apparel/purchase_order_details.html', {
        'purchase_order': purchase_order,
        'po_details': po_details
    })


# Sales Orders List
def sales_orders(request):
    query = request.GET.get('q', '')
    orders = SalesOrder.objects.all().order_by('-sales_id')
    if query:
        orders = orders.filter(customer__full_name__icontains=query)
    return render(request, 'apparel/sales_orders_list.html', {
        'sales_orders': orders,
        'query': query,
    })


def shipping_list(request):
    query = request.GET.get('q', '')
    shippings = ShippingDetails.objects.select_related('sales_order').all()
    if query:
        shippings = shippings.filter(recipient_name__icontains=query)
    return render(request, 'apparel/shipping_details.html', {'shippings': shippings, 'query': query})


# Create Sales Order
def create_sales_order(request):
    import json

    if request.method == 'POST':
        customer_id = request.POST.get('customer')
        type_of_sale = request.POST.get('type_of_sale')
        payment_method = request.POST.get('payment_method')
        discount = request.POST.get('discount') or 0
        freight_cost = request.POST.get('freight_cost') or 0  # NEW

        customer = Customer.objects.get(id=customer_id)
        employee_id = request.session.get('employee_id')
        employee = Employee.objects.get(id=employee_id) if employee_id else None

        sales_order = SalesOrder.objects.create(
            customer=customer,
            employee=employee,
            type_of_sale=type_of_sale,
            payment_method=payment_method,
            discount=discount,
            freight_cost=freight_cost,  # NEW
            total_amount=0,
        )

        product_ids = request.POST.getlist('product[]')
        quantities = request.POST.getlist('quantity[]')
        unit_prices = request.POST.getlist('unit_price[]')
        new_prices = request.POST.getlist('new_price[]')
        sd_discounts = request.POST.getlist('sd_discount[]')

        total = 0

        for i in range(len(product_ids)):
            product = Product.objects.get(id=product_ids[i])
            quantity = int(quantities[i])

            if new_prices[i]:
                unit_price = float(new_prices[i])
            else:
                unit_price = float(unit_prices[i])

            sd_discount = float(sd_discounts[i]) if sd_discounts[i] else 0

            subtotal = quantity * unit_price
            discount_amount = subtotal * (sd_discount / 100)
            subtotal -= discount_amount
            total += subtotal

            SalesOrderDetails.objects.create(
                sales_order=sales_order,
                product=product,
                quantity_sold=quantity,
                unit_price=unit_price,
                sd_discount=sd_discount,
                subtotal=round(subtotal, 2),
            )

        order_discount = float(discount)
        total_discount = total * (order_discount / 100)
        final_total = total - total_discount + float(freight_cost)  # include freight

        sales_order.total_amount = round(final_total, 2)
        sales_order.save()

        if type_of_sale == 'Online':
            from .models import ShippingDetails
            ShippingDetails.objects.create(
                sales_order=sales_order,
                recipient_name=request.POST.get('recipient_name', ''),
                contact_number=request.POST.get('contact_number', ''),
                shipping_address=request.POST.get('shipping_address', ''),
                shipping_method=request.POST.get('shipping_method', 'LBC'),
                shipping_fee=request.POST.get('shipping_fee') or 0,
                shipping_status='Pending',
            )
        else:
            Invoice.objects.create(
                ref=f'INV-{sales_order.sales_id}',
                date=sales_order.sales_date,
                customer=customer.full_name,
                description=f'Sales Order SO-{sales_order.sales_id} (Walk-in)',
                amount=sales_order.total_amount,
                status='paid',
                sales_order=sales_order,
            )

        audit(request, 'CREATE', f'{request.session.get("employee_name")} created Sales Order SO-{sales_order.sales_id} for customer "{customer.full_name}" (₱{sales_order.total_amount}).')
        return redirect('sales_orders')

    # GET
    customers = Customer.objects.all()
    categories = Category.objects.all()
    brands = Brand.objects.select_related('category').all()
    products_qs = Product.objects.prefetch_related('sizes', 'colors').all()

    product_data = {}
    for p in products_qs:
        product_data[p.id] = {
            'name': p.product_name,
            'brand_id': p.brand_id,
            'category_id': p.category_id,   # include category_id so JS can filter
            'price': float(p.selling_price),
            'colors': [{'id': c.id, 'name': c.color_name} for c in p.colors.all()],
            'sizes': [{'id': s.id, 'name': s.size_name} for s in p.sizes.all()],
        }

    return render(request, 'apparel/create_sales_order.html', {
        'customers': customers,
        'products': products_qs,
        'categories': categories,
        'brands': brands,
        'product_data_json': json.dumps(product_data),
        'employee_name': request.session.get('employee_name'),
        'employee_id': request.session.get('employee_id'),
    })


# Sales Order Details
def sales_order_details(request, pk):
    sales_order = get_object_or_404(SalesOrder, pk=pk)
    so_details = SalesOrderDetails.objects.filter(sales_order=sales_order).select_related('product')
    return render(request, 'apparel/sales_order_details.html', {
        'sales_order': sales_order,
        'so_details': so_details,
    })


def accounting(request):
    import json
    from datetime import date
    import calendar

    today = date.today()
    months_labels = []
    monthly_income = []
    monthly_expenses = []

    for i in range(5, -1, -1):
        total_months = today.month - 1 - i
        year_num = today.year + total_months // 12
        month_num = total_months % 12 + 1
        if month_num <= 0:
            month_num += 12
            year_num -= 1

        income = Invoice.objects.filter(
            date__year=year_num, date__month=month_num, status='paid'
        ).aggregate(total=Sum('amount'))['total'] or 0

        expense = Expense.objects.filter(
            date__year=year_num, date__month=month_num
        ).aggregate(total=Sum('amount'))['total'] or 0

        months_labels.append(calendar.month_abbr[month_num])
        monthly_income.append(float(income))
        monthly_expenses.append(float(expense))

    return render(request, 'apparel/accounting.html', {
        'chart_labels_json': json.dumps(months_labels),
        'chart_income_json': json.dumps(monthly_income),
        'chart_expenses_json': json.dumps(monthly_expenses),
        'employee_name': request.session.get('employee_name'),
        'employee_role': request.session.get('employee_role'),
    })

def api_invoices(request):
    from .models import Invoice
    import json
    from django.http import JsonResponse
    if request.method == 'GET':
        data = list(Invoice.objects.all().order_by('-id').values())
        return JsonResponse(data, safe=False)
    if request.method == 'POST':
        body = json.loads(request.body)
        inv = Invoice.objects.create(
            ref=body.get('ref',''),
            date=body.get('date'),
            customer=body.get('customer',''),
            description=body.get('description',''),
            amount=body.get('amount',0),
            status=body.get('status','pending'),
        )
        audit(request, 'CREATE', f'{request.session.get("employee_name")} created invoice {inv.ref} for {inv.customer} (₱{inv.amount}).')
        return JsonResponse({'id':inv.id,'ref':inv.ref,'date':str(inv.date),'customer':inv.customer,'description':inv.description,'amount':float(inv.amount),'status':inv.status})

def api_invoice_detail(request, pk):
    from .models import Invoice
    from django.http import JsonResponse
    inv = get_object_or_404(Invoice, pk=pk)
    if request.method == 'DELETE':
        audit(request, 'DELETE', f'{request.session.get("employee_name")} deleted invoice {inv.ref} (ID: {pk}).')
        inv.delete()
        return JsonResponse({}, status=204)

def api_receipts(request):
    from .models import Receipt
    import json
    from django.http import JsonResponse
    if request.method == 'GET':
        data = list(Receipt.objects.all().order_by('-id').values())
        return JsonResponse(data, safe=False)
    if request.method == 'POST':
        body = json.loads(request.body)
        rec = Receipt.objects.create(
            ref=body.get('ref',''),
            date=body.get('date'),
            source=body.get('source',''),
            description=body.get('description',''),
            amount=body.get('amount',0),
            receipt_type=body.get('receipt_type','purchase'),
        )
        audit(request, 'CREATE', f'{request.session.get("employee_name")} created receipt {rec.ref} from {rec.source} (₱{rec.amount}).')
        return JsonResponse({'id':rec.id,'ref':rec.ref,'date':str(rec.date),'source':rec.source,'description':rec.description,'amount':float(rec.amount),'receipt_type':rec.receipt_type})

def api_receipt_detail(request, pk):
    from .models import Receipt
    from django.http import JsonResponse
    rec = get_object_or_404(Receipt, pk=pk)
    if request.method == 'DELETE':
        audit(request, 'DELETE', f'{request.session.get("employee_name")} deleted receipt {rec.ref} (ID: {pk}).')
        rec.delete()
        return JsonResponse({}, status=204)

def api_expenses(request):
    from .models import Expense
    import json
    from django.http import JsonResponse
    if request.method == 'GET':
        data = list(Expense.objects.all().order_by('-id').values())
        return JsonResponse(data, safe=False)
    if request.method == 'POST':
        body = json.loads(request.body)
        exp = Expense.objects.create(
            date=body.get('date'),
            category=body.get('category','Other'),
            description=body.get('description',''),
            amount=body.get('amount',0),
            payment_method=body.get('payment_method','Cash'),
        )
        audit(request, 'CREATE', f'{request.session.get("employee_name")} created expense "{exp.category}" on {exp.date} (₱{exp.amount}).')
        return JsonResponse({'id':exp.id,'date':str(exp.date),'category':exp.category,'description':exp.description,'amount':float(exp.amount),'payment_method':exp.payment_method})

def api_expense_detail(request, pk):
    from .models import Expense
    from django.http import JsonResponse
    exp = get_object_or_404(Expense, pk=pk)
    if request.method == 'DELETE':
        audit(request, 'DELETE', f'{request.session.get("employee_name")} deleted expense ID {pk} ({exp.category}, ₱{exp.amount}).')
        exp.delete()
        return JsonResponse({}, status=204)

# JSON endpoint for sales orders (accounting)
def api_sales_orders_accounting(request):
    from .models import SalesOrder
    from django.http import JsonResponse
    if request.method == 'GET':
        data = list(
            SalesOrder.objects.all().order_by('-sales_id').values(
                'sales_id', 'customer__full_name', 'employee__full_name',
                'total_amount', 'type_of_sale', 'payment_method', 'sales_date'
            )
        )
        return JsonResponse(data, safe=False)

# JSON endpoint for purchase orders (accounting)
def api_purchase_orders_accounting(request):
    from .models import PurchaseOrder
    from django.http import JsonResponse
    if request.method == 'GET':
        data = list(
            PurchaseOrder.objects.all().order_by('-purchase_order_id').values(
                'purchase_order_id', 'supplier__supplier_name', 'employee__full_name',
                'total_cost', 'po_discount', 'status', 'order_date', 'expected_delivery_date'
            )
        )
        return JsonResponse(data, safe=False)


# ── EXCEL EXPORT ENDPOINTS ──────────────────────────────────────────────────
def export_invoices_excel(request):
    import openpyxl
    from django.http import HttpResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Invoices'
    headers = ['ID', 'Ref', 'Date', 'Customer', 'Description', 'Amount', 'Status', 'Created At']
    ws.append(headers)
    for inv in Invoice.objects.all().order_by('-id'):
        ws.append([inv.id, inv.ref, str(inv.date), inv.customer,
                   inv.description or '', float(inv.amount), inv.status, str(inv.created_at)])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'
    wb.save(response)
    return response


def export_receipts_excel(request):
    import openpyxl
    from django.http import HttpResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Receipts'
    headers = ['ID', 'Ref', 'Date', 'Source', 'Description', 'Amount', 'Type', 'Created At']
    ws.append(headers)
    for rec in Receipt.objects.all().order_by('-id'):
        ws.append([rec.id, rec.ref, str(rec.date), rec.source,
                   rec.description or '', float(rec.amount), rec.receipt_type, str(rec.created_at)])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="receipts.xlsx"'
    wb.save(response)
    return response


def export_expenses_excel(request):
    import openpyxl
    from django.http import HttpResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'
    headers = ['ID', 'Date', 'Category', 'Description', 'Amount', 'Payment Method', 'Created At']
    ws.append(headers)
    for exp in Expense.objects.all().order_by('-id'):
        ws.append([exp.id, str(exp.date), exp.category,
                   exp.description or '', float(exp.amount), exp.payment_method, str(exp.created_at)])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="expenses.xlsx"'
    wb.save(response)
    return response


def export_sales_orders_excel(request):
    import openpyxl
    from django.http import HttpResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Orders'
    headers = ['SO ID', 'Date', 'Customer', 'Employee', 'Type', 'Payment', 'Discount %', 'Freight', 'Total Amount']
    ws.append(headers)
    for so in SalesOrder.objects.select_related('customer', 'employee').order_by('-sales_id'):
        ws.append([
            f'SO-{so.sales_id}', str(so.sales_date),
            so.customer.full_name if so.customer else '',
            so.employee.full_name if so.employee else '',
            so.type_of_sale, so.payment_method,
            float(so.discount), float(so.freight_cost), float(so.total_amount),
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sales_orders.xlsx"'
    wb.save(response)
    return response


def export_purchase_orders_excel(request):
    import openpyxl
    from django.http import HttpResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Purchase Orders'
    headers = ['PO ID', 'Order Date', 'Supplier', 'Employee', 'Discount %', 'Total Cost', 'Status', 'Expected Delivery']
    ws.append(headers)
    for po in PurchaseOrder.objects.select_related('supplier', 'employee').order_by('-purchase_order_id'):
        ws.append([
            f'PO-{po.purchase_order_id}', str(po.order_date),
            po.supplier.supplier_name if po.supplier else '',
            po.employee.full_name if po.employee else '',
            float(po.po_discount), float(po.total_cost),
            po.status, str(po.expected_delivery_date),
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="purchase_orders.xlsx"'
    wb.save(response)
    return response

def audit_log_view(request):
    employee_id = request.session.get('employee_id')
    if not employee_id:
        return redirect('login')

    role = request.session.get('employee_role')
    if role != 'Admin':
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Access denied.")

    from django.db import connection
    connection.close()  # Force fresh DB connection — prevents stale/cached query results

    logs = list(AuditLog.objects.select_related('employee').order_by('-datetime'))
    response = render(request, 'apparel/audit_log.html', {
        'logs': logs,
        'employee_name': request.session.get('employee_name'),
        'employee_role': role,
    })
    # Prevent browser and proxy from caching this page
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

def update_shipping_status(request, pk):
    from .models import ShippingDetails
    shipping = get_object_or_404(ShippingDetails, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('shipping_status', 'Pending')
        shipping.shipping_status = new_status
        shipping.save()

        if new_status == 'Delivered':
            sales_order = shipping.sales_order
            try:
                invoice = sales_order.invoice
                invoice.status = 'paid'
                invoice.save()
            except Exception:
                from django.utils import timezone as tz
                Invoice.objects.create(
                    ref=f'INV-{sales_order.sales_id}',
                    date=tz.now().date(),
                    customer=sales_order.customer.full_name,
                    description=f'Sales Order SO-{sales_order.sales_id} (Online) — Delivered',
                    amount=sales_order.total_amount,
                    status='paid',
                    sales_order=sales_order,
                )

        audit(request, 'STATUS', f'{request.session.get("employee_name")} updated shipping SHIP-{pk} status to "{new_status}".')

    return redirect('shipping_list')


def update_shipping_from_so(request, pk):
    sales_order = get_object_or_404(SalesOrder, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('shipping_status', 'Pending')
        try:
            shipping = sales_order.shipping
            shipping.shipping_status = new_status
            shipping.save()

            if new_status == 'Delivered':
                try:
                    invoice = sales_order.invoice
                    invoice.status = 'paid'
                    invoice.save()
                except Exception:
                    from django.utils import timezone as tz
                    Invoice.objects.create(
                        ref=f'INV-{sales_order.sales_id}',
                        date=tz.now().date(),
                        customer=sales_order.customer.full_name,
                        description=f'Sales Order SO-{sales_order.sales_id} (Online) — Delivered',
                        amount=sales_order.total_amount,
                        status='paid',
                        sales_order=sales_order,
                    )
        except Exception:
            pass

        audit(request, 'STATUS', f'{request.session.get("employee_name")} updated shipping status for SO-{pk} to "{new_status}".')

    return redirect('sales_order_details', pk=pk)

def transactions_page(request):
    product_id = request.GET.get('product')
    search = request.GET.get('search', '').strip()
    sort = request.GET.get('sort', 'newest').strip()

    if product_id:
        try:
            product_id = int(product_id)
        except ValueError:
            product_id = None
    else:
        product_id = None

    purchase_qs = PurchaseOrderDetails.objects.select_related(
        'purchase_order', 'purchase_order__supplier', 'product'
    ).filter(purchase_order__status='Received')

    sales_qs = SalesOrderDetails.objects.select_related(
        'sales_order', 'sales_order__customer', 'product'
    )

    if product_id is not None:
        purchase_qs = purchase_qs.filter(product_id=product_id)
        sales_qs = sales_qs.filter(product_id=product_id)

    if search:
        purchase_qs = purchase_qs.filter(product__product_name__icontains=search)
        sales_qs = sales_qs.filter(product__product_name__icontains=search)

    transactions = []

    for item in purchase_qs:
        transactions.append({
            'date': item.purchase_order.order_date,
            'type': 'PURCHASE',
            'reference': f"PO-{item.purchase_order.purchase_order_id}",
            'product_id': item.product.id,
            'product_name': item.product.product_name,
            'partner': item.purchase_order.supplier.supplier_name,
            'qty_in': item.quantity_ordered,
            'qty_out': 0,
            'net': item.quantity_ordered,
            'unit_price': item.unit_cost,
            'status': item.purchase_order.status,
        })

    for item in sales_qs:
        transactions.append({
            'date': item.sales_order.sales_date,
            'type': 'SALE',
            'reference': f"SO-{item.sales_order.sales_id}",
            'product_id': item.product.id,
            'product_name': item.product.product_name,
            'partner': item.sales_order.customer.full_name,
            'qty_in': 0,
            'qty_out': item.quantity_sold,
            'net': -item.quantity_sold,
            'unit_price': item.unit_price,
            'status': 'Completed',
        })

    if sort == 'oldest':
        transactions.sort(key=lambda x: (x['date'], x['reference']))
    else:
        transactions.sort(key=lambda x: (x['date'], x['reference']), reverse=True)

    balances = {}
    oldest_first = sorted(transactions, key=lambda x: (x['date'], x['reference']))

    for tx in oldest_first:
        pid = tx['product_id']
        balances[pid] = balances.get(pid, 0) + tx['net']
        tx['_computed_balance'] = balances[pid]

    balance_lookup = {
        (tx['product_id'], tx['reference'], tx['type'], tx['date']): tx['_computed_balance']
        for tx in oldest_first
    }

    for tx in transactions:
        tx['balance'] = balance_lookup[
            (tx['product_id'], tx['reference'], tx['type'], tx['date'])
        ]

    products = Product.objects.all().order_by('product_name')

    if search:
        suggestions = list(
            Product.objects.filter(product_name__icontains=search)
            .order_by('product_name')
            .values_list('product_name', flat=True)
            .distinct()[:8]
        )
    else:
        suggestions = list(
            Product.objects.order_by('product_name')
            .values_list('product_name', flat=True)[:8]
        )

    selected_product_name = ''
    if product_id:
        selected = products.filter(id=product_id).first()
        if selected:
            selected_product_name = selected.product_name

    return render(request, 'apparel/transactions.html', {
        'transactions': transactions,
        'products': products,
        'selected_product': product_id,
        'selected_product_name': selected_product_name,
        'employee_name': request.session.get('employee_name'),
        'employee_role': request.session.get('employee_role'),
        'search': search,
        'sort': sort,
        'suggestions': suggestions,
    })

def product_transaction_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    purchase_qs = PurchaseOrderDetails.objects.select_related(
        'purchase_order', 'purchase_order__supplier'
    ).filter(
        product_id=product_id,
        purchase_order__status='Received'
    ).order_by('-purchase_order__order_date')

    sales_qs = SalesOrderDetails.objects.select_related(
        'sales_order', 'sales_order__customer'
    ).filter(
        product_id=product_id
    ).order_by('-sales_order__sales_date')

    purchases = []
    total_purchased = 0
    for item in purchase_qs:
        purchases.append({
            'date': str(item.purchase_order.order_date),
            'reference': f"PO-{item.purchase_order.purchase_order_id}",
            'partner': item.purchase_order.supplier.supplier_name,
            'qty_in': item.quantity_ordered,
            'unit_price': str(item.unit_cost),
            'status': item.purchase_order.status,
        })
        total_purchased += item.quantity_ordered

    sales = []
    total_sold = 0
    for item in sales_qs:
        sales.append({
            'date': str(item.sales_order.sales_date),
            'reference': f"SO-{item.sales_order.sales_id}",
            'partner': item.sales_order.customer.full_name,
            'qty_out': item.quantity_sold,
            'unit_price': str(item.unit_price),
            'status': 'Completed',
        })
        total_sold += item.quantity_sold

    current_balance = total_purchased - total_sold

    return JsonResponse({
        'product_code': product.product_code,
        'current_balance': current_balance,
        'total_purchased': total_purchased,
        'total_sold': total_sold,
        'purchases': purchases,
        'sales': sales,
    })